"""Smoke tests for skills/qc-questions/scripts/qc_xlsx.py.

Goal: every behaviour documented in the skill is locked in by a test so
regressions surface fast. New skill prescriptions follow the writing-skills
Iron Law: add a failing test here BEFORE you change skill files.
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = REPO_ROOT / "skills" / "qc-questions" / "scripts" / "qc_xlsx.py"
SAMPLE_XLSX = REPO_ROOT / "examples" / "sample_bank.xlsx"
EXPECTED_VERDICTS = REPO_ROOT / "examples" / "sample_bank.expected_verdicts.json"

GREEN = "00C6EFCE"
AMBER = "00FFEB9C"
DARK_AMBER = "00FFD966"
RED = "00FFC7CE"

sys.path.insert(0, str(REPO_ROOT / "skills" / "qc-questions" / "scripts"))

import qc_xlsx as qc  # noqa: E402  -- import follows sys.path mutation above

# ---------------------------------------------------------------------------
# Pure helpers
# ---------------------------------------------------------------------------

def test_strip_html_unwraps_p_tag():
    assert qc.strip_html("<p>Hello world</p>") == "Hello world"


def test_strip_html_collapses_whitespace_and_handles_br():
    out = qc.strip_html("<p>line one<br>line two</p>")
    assert out == "line one\nline two"


def test_strip_html_passthrough_for_plain_text():
    assert qc.strip_html("plain text") == "plain text"


def test_strip_html_handles_none():
    assert qc.strip_html(None) == ""


def test_maybe_html_wrap_wraps_plain_text_for_content_field():
    assert qc._maybe_html_wrap("content", "abc") == "<p>abc</p>"


def test_maybe_html_wrap_leaves_existing_html_alone():
    assert qc._maybe_html_wrap("content", "<p>abc</p>") == "<p>abc</p>"


def test_maybe_html_wrap_does_not_wrap_correctOption_or_difficulty():
    assert qc._maybe_html_wrap("correctOption", "option2") == "option2"
    assert qc._maybe_html_wrap("difficulty", "MEDIUM") == "MEDIUM"


def test_normalize_field_aliases_stem_to_content():
    assert qc._normalize_field("stem") == "content"
    assert qc._normalize_field("option3") == "option3"


# ---------------------------------------------------------------------------
# `read` subcommand — HTML stripping + key holdback
# ---------------------------------------------------------------------------

def _run_read():
    proc = subprocess.run(
        [sys.executable, str(SCRIPT), "read", str(SAMPLE_XLSX)],
        capture_output=True,
        text=True,
        check=True,
    )
    return json.loads(proc.stdout)


def test_read_separates_keys_from_questions():
    parsed = _run_read()
    assert "questions" in parsed
    assert "_keys" in parsed
    assert len(parsed["questions"]) == 9
    assert len(parsed["_keys"]) == 9
    for q in parsed["questions"]:
        assert "correctOption" not in q, "correctOption must not leak into question payload"


def test_read_holds_back_both_audit_targets():
    """Dual-blind invariant: both correctOption AND marked difficulty must be
    stripped from `questions` and held in `_keys`. The marked difficulty is the
    spec the QC is auditing against — if a subagent saw it, the difficulty
    rating would be a post-hoc rationalization of the tag, not an audit signal.
    """
    parsed = _run_read()
    for q in parsed["questions"]:
        assert "correctOption" not in q, "correctOption leaked into questions"
        assert "difficulty" not in q, "marked difficulty leaked into questions"
    for k in parsed["_keys"]:
        assert "correctOption" in k, "correctOption missing from _keys"
        assert "difficulty" in k, "marked difficulty missing from _keys"
        assert set(k.keys()) >= {"row", "correctOption", "difficulty"}


def test_read_strips_html_from_content_and_options():
    parsed = _run_read()
    first = parsed["questions"][0]
    assert first["content"] == "Identify the grammatically correct sentence."
    assert "<p>" not in first["option1"]


def test_read_preserves_row_numbers_starting_at_2():
    parsed = _run_read()
    rows = [q["row"] for q in parsed["questions"]]
    assert rows == [2, 3, 4, 5, 6, 7, 8, 9, 10]


# ---------------------------------------------------------------------------
# `write` subcommand — full end-to-end against the sample bank
# ---------------------------------------------------------------------------

@pytest.fixture()
def written_workbook(tmp_path: Path) -> Path:
    out = tmp_path / "out.xlsx"
    subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "write",
            str(SAMPLE_XLSX),
            str(EXPECTED_VERDICTS),
            str(out),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    return out


def test_write_creates_four_sheets_in_order(written_workbook: Path):
    """v0.4.0: QC Summary sheet is now the first sheet."""
    wb = openpyxl.load_workbook(written_workbook)
    assert wb.sheetnames == ["QC Summary", "Sheet1", "QC Legend", "Corrected"]


def test_qc_summary_sheet_has_bank_counts(written_workbook: Path):
    """v0.4.0: QC Summary sheet renders bank-level counts at the top."""
    wb = openpyxl.load_workbook(written_workbook)
    s = wb["QC Summary"]
    text = "\n".join(
        str(cell.value) for row in s.iter_rows() for cell in row if cell.value is not None
    )
    assert "Bank counts" in text
    assert "total=9" in text


def test_blueprint_review_fires_on_same_direction_two_band_cluster(tmp_path: Path):
    """v0.4.0: when >20% of items in a section show same-direction 2-band gaps,
    the blueprint_review observation must fire on the QC Summary sheet.

    Synthesize a 3-row bank where all three items are marked HARD but blind-rate
    EASY (2-band overrated). Run the writer and assert the QC Summary sheet
    carries 'blueprint_review' and the section name.
    """
    # Build a tiny 3-row xlsx
    bank = tmp_path / "blueprint_test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "content", "option1", "option2", "option3", "option4",
        "option5", "option6", "correctOption", "questionType",
        "subject", "topics", "difficulty",
    ]
    ws.append(headers)
    for i in range(3):
        ws.append([
            f"Q{i+1} stem text", "A", "B", "C", "D", None, None,
            "option1", "MULTIPLE_CHOICE_QUESTION",
            "Test Subject", "Test Topic", "HARD",
        ])
    wb.save(bank)

    # Verdicts: all three blind-rated EASY against marked HARD (2-band over)
    verdicts = tmp_path / "v.json"
    verdicts.write_text(json.dumps([
        {"row": r, "status": "NEEDS_EDITS", "correctness_issue": None,
         "difficulty_issue": f"row {r} blind EASY vs marked HARD",
         "edits": [], "confidence": "LOW",
         "my_blind_difficulty": "EASY"}
        for r in (2, 3, 4)
    ]))

    out = tmp_path / "out.xlsx"
    proc = subprocess.run(
        [sys.executable, str(SCRIPT), "write", str(bank), str(verdicts), str(out)],
        check=True, capture_output=True, text=True,
    )
    assert "QC Summary" in proc.stderr

    wb_out = openpyxl.load_workbook(out)
    assert "QC Summary" in wb_out.sheetnames
    s = wb_out["QC Summary"]
    text = "\n".join(
        str(cell.value) for row in s.iter_rows() for cell in row if cell.value is not None
    )
    assert "blueprint_review" in text, "blueprint_review observation did not fire on a 100% same-direction 2-band cluster"
    assert "Test Subject / Test Topic" in text, "section name missing from blueprint_review observation"
    assert "OVERRATED" in text or "overrated" in text, "direction missing from blueprint_review observation"


def test_qc_status_lands_at_next_truly_empty_column(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    # The sample bank has 12 named headers, so qc_status should land at col 13.
    assert "qc_status" in headers
    assert headers.index("qc_status") == 12


def test_qc_changes_lands_immediately_after_qc_status(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    assert "qc_changes" in headers
    assert headers.index("qc_changes") == headers.index("qc_status") + 1


def test_qc_changes_column_has_wide_width(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(headers.index("qc_changes") + 1)
    assert ws.column_dimensions[col_letter].width == 80


def test_qc_status_fills_match_severity(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ws = wb["Sheet1"]
    col = [c.value for c in ws[1]].index("qc_status") + 1
    expected = {
        2: ("ALIGNED", GREEN),
        3: ("ALIGNED", GREEN),
        4: ("NEEDS_EDITS", AMBER),
        5: ("NEEDS_EDITS", AMBER),
        6: ("ALIGNED", GREEN),
        7: ("NEEDS_EDITS", AMBER),
        8: ("NEEDS_EDITS", RED),   # LOW confidence escalation
        9: ("ALIGNED", GREEN),
    }
    for row_no, (status, fill) in expected.items():
        cell = ws.cell(row=row_no, column=col)
        assert cell.value == status, f"row {row_no} status"
        assert cell.fill.start_color.rgb == fill, f"row {row_no} fill"


def test_qc_changes_aligned_rows_are_empty(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ws = wb["Sheet1"]
    col = [c.value for c in ws[1]].index("qc_changes") + 1
    for row_no in (2, 3, 6, 9):
        cell = ws.cell(row=row_no, column=col)
        assert cell.value is None, f"row {row_no} qc_changes should be empty"


def test_qc_changes_needs_edits_row_carries_full_narrative(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ws = wb["Sheet1"]
    col = [c.value for c in ws[1]].index("qc_changes") + 1
    # Row 4: wrong-key correctness issue, no difficulty issue, one edit.
    cell = ws.cell(row=4, column=col)
    val = cell.value
    assert val is not None
    assert "CORRECTNESS:" in val
    assert "DIFFICULTY:" not in val  # difficulty_issue is null on row 4
    assert "EDITS APPLIED:" in val
    assert "  • correctOption:" in val
    assert "fixes correctness" in val
    assert cell.fill.start_color.rgb == AMBER
    assert cell.alignment.wrap_text is True


def test_qc_changes_difficulty_only_row_has_difficulty_section(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ws = wb["Sheet1"]
    col = [c.value for c in ws[1]].index("qc_changes") + 1
    # Row 5: difficulty mismatch only, one option2 edit.
    val = ws.cell(row=5, column=col).value
    assert "CORRECTNESS:" not in val
    assert "DIFFICULTY:" in val
    assert "EDITS APPLIED:" in val
    assert "  • option2:" in val


def test_qc_changes_low_confidence_no_edits_emits_human_review_line(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ws = wb["Sheet1"]
    col = [c.value for c in ws[1]].index("qc_changes") + 1
    # Row 8: construct mismatch, LOW confidence, zero edits.
    cell = ws.cell(row=8, column=col)
    val = cell.value
    assert "CORRECTNESS:" in val
    assert "EDITS: none auto-applied — human review required" in val
    assert cell.fill.start_color.rgb == RED


def test_qc_changes_sections_separated_by_blank_line(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ws = wb["Sheet1"]
    col = [c.value for c in ws[1]].index("qc_changes") + 1
    # Row 4 has CORRECTNESS + EDITS APPLIED — sections must be blank-line-separated.
    val = ws.cell(row=4, column=col).value
    assert "\n\n" in val


def test_corrected_aligned_rows_are_populated_verbatim_and_green(written_workbook: Path):
    """ALIGNED rows must carry the original content so the Corrected sheet is
    upload-ready as the single source of truth. The whole row is green-filled."""
    src = openpyxl.load_workbook(SAMPLE_XLSX)
    src_ws = src["Sheet1"]
    src_headers = [c.value for c in src_ws[1]]
    wb = openpyxl.load_workbook(written_workbook)
    cs = wb["Corrected"]
    dst_headers = [c.value for c in cs[1]]
    # Audit columns must NOT appear on the Corrected sheet.
    assert "qc_status" not in dst_headers
    assert "qc_changes" not in dst_headers
    for row_no in (2, 3, 6, 9):
        for h in dst_headers:
            src_col = src_headers.index(h) + 1
            dst_col = dst_headers.index(h) + 1
            src_val = src_ws.cell(row=row_no, column=src_col).value
            cell = cs.cell(row=row_no, column=dst_col)
            assert cell.value == src_val, (
                f"row {row_no} col {h} should equal source verbatim: "
                f"{src_val!r} vs {cell.value!r}"
            )
            assert cell.fill.start_color.rgb == GREEN, (
                f"row {row_no} col {h} should be green-filled"
            )
            # Not edited → no dark-amber/bold accent.
            assert cell.font.bold is False


def test_corrected_needs_edits_row_has_post_edit_content(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    cs = wb["Corrected"]
    headers = [c.value for c in cs[1]]
    co_col = headers.index("correctOption") + 1
    # Row 4 should have its correctOption flipped from option3 to option2.
    cell = cs.cell(row=4, column=co_col)
    assert cell.value == "option2"
    assert cell.font.bold is True
    assert cell.fill.start_color.rgb == DARK_AMBER


def test_corrected_unedited_cells_in_needs_edits_row_are_amber_not_dark_amber(
    written_workbook: Path,
):
    wb = openpyxl.load_workbook(written_workbook)
    cs = wb["Corrected"]
    headers = [c.value for c in cs[1]]
    content_col = headers.index("content") + 1
    # Row 4: content was not edited; should be amber, not dark amber.
    cell = cs.cell(row=4, column=content_col)
    assert cell.fill.start_color.rgb == AMBER
    assert cell.font.bold is False


def test_corrected_low_confidence_row_is_red(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    cs = wb["Corrected"]
    # Row 8 is the construct-mismatch escalation (LOW confidence, no edits).
    cell = cs.cell(row=8, column=1)
    assert cell.fill.start_color.rgb == RED


def test_immutable_fields_are_preserved_in_corrected_sheet(written_workbook: Path):
    """`subject`, `topics`, `difficulty` must never be silently changed.

    Now that ALIGNED rows are populated verbatim, every row's immutable cells
    must match the source — no None passthrough.
    """
    src = openpyxl.load_workbook(SAMPLE_XLSX)
    dst = openpyxl.load_workbook(written_workbook)
    src_ws = src["Sheet1"]
    dst_cs = dst["Corrected"]
    src_headers = [c.value for c in src_ws[1]]
    dst_headers = [c.value for c in dst_cs[1]]
    for field in ("subject", "topics", "difficulty"):
        src_col = src_headers.index(field) + 1
        dst_col = dst_headers.index(field) + 1
        for row_no in range(2, 10):
            src_val = src_ws.cell(row=row_no, column=src_col).value
            dst_val = dst_cs.cell(row=row_no, column=dst_col).value
            assert dst_val == src_val, (
                f"immutable {field} changed at row {row_no}: {src_val!r} -> {dst_val!r}"
            )


# ---------------------------------------------------------------------------
# Immutable-field rejection + --allow-retag override
# ---------------------------------------------------------------------------

def _retag_verdicts(tmp_path: Path) -> Path:
    v = [
        {
            "row": 2,
            "status": "NEEDS_EDITS",
            "correctness_issue": None,
            "difficulty_issue": "test",
            "edits": [
                {"field": "difficulty", "from": "EASY", "to": "MEDIUM", "why": "test"},
                {"field": "subject", "from": "Verbal Ability", "to": "Logical", "why": "test"},
            ],
            "confidence": "HIGH",
        }
    ]
    p = tmp_path / "retag.json"
    p.write_text(json.dumps(v))
    return p


def test_immutable_field_edits_dropped_by_default(tmp_path: Path):
    verdicts = _retag_verdicts(tmp_path)
    out = tmp_path / "out.xlsx"
    proc = subprocess.run(
        [sys.executable, str(SCRIPT), "write", str(SAMPLE_XLSX), str(verdicts), str(out)],
        capture_output=True,
        text=True,
        check=True,
    )
    # The writer should warn on stderr.
    assert "immutable field 'difficulty'" in proc.stderr
    assert "immutable field 'subject'" in proc.stderr
    # And the Corrected sheet should retain the marked tags.
    wb = openpyxl.load_workbook(out)
    cs = wb["Corrected"]
    headers = [c.value for c in cs[1]]
    diff_col = headers.index("difficulty") + 1
    subj_col = headers.index("subject") + 1
    assert cs.cell(row=2, column=diff_col).value == "EASY"
    assert cs.cell(row=2, column=subj_col).value == "Verbal Ability"


def test_allow_retag_flag_applies_immutable_field_edits(tmp_path: Path):
    verdicts = _retag_verdicts(tmp_path)
    out = tmp_path / "out.xlsx"
    subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "write",
            str(SAMPLE_XLSX),
            str(verdicts),
            str(out),
            "--allow-retag",
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    wb = openpyxl.load_workbook(out)
    cs = wb["Corrected"]
    headers = [c.value for c in cs[1]]
    diff_col = headers.index("difficulty") + 1
    subj_col = headers.index("subject") + 1
    # With the flag, retags ARE applied. Cells should be dark-amber + bold.
    diff_cell = cs.cell(row=2, column=diff_col)
    subj_cell = cs.cell(row=2, column=subj_col)
    assert diff_cell.value == "MEDIUM"
    assert subj_cell.value == "Logical"
    assert diff_cell.font.bold is True
    assert subj_cell.font.bold is True


# ---------------------------------------------------------------------------
# Auto-LOW guard: NEEDS_EDITS at MED/HIGH with empty edits is a regression.
# The writer warns + bumps the row's fill to RED + emits the human-review
# narrative, so the gap is visible to reviewers without breaking the pipeline.
# ---------------------------------------------------------------------------

def _empty_edits_verdict(tmp_path: Path, *, confidence: str) -> Path:
    v = [
        {
            "row": 2,
            "status": "NEEDS_EDITS",
            "correctness_issue": "fake correctness issue for guard test",
            "difficulty_issue": None,
            "edits": [],
            "confidence": confidence,
        }
    ]
    p = tmp_path / f"empty_{confidence}.json"
    p.write_text(json.dumps(v))
    return p


def test_writer_warns_and_reds_empty_edits_at_high(tmp_path: Path):
    """A NEEDS_EDITS verdict at HIGH confidence with empty edits is a subagent
    regression. Writer must (1) warn on stderr with the canonical phrasing,
    (2) bump qc_status + qc_changes + Corrected row fills to RED, (3) emit
    the 'human review required' narrative."""
    verdicts = _empty_edits_verdict(tmp_path, confidence="HIGH")
    out = tmp_path / "out.xlsx"
    proc = subprocess.run(
        [sys.executable, str(SCRIPT), "write", str(SAMPLE_XLSX), str(verdicts), str(out)],
        capture_output=True, text=True, check=True,
    )
    assert (
        "row 2: NEEDS_EDITS at confidence HIGH but no edits supplied"
        in proc.stderr
    ), proc.stderr
    assert "Treating as LOW for fill colour" in proc.stderr

    wb = openpyxl.load_workbook(out)
    ws = wb["Sheet1"]
    headers = [c.value for c in ws[1]]
    qs = headers.index("qc_status") + 1
    qc = headers.index("qc_changes") + 1
    assert ws.cell(row=2, column=qs).fill.start_color.rgb == RED
    qc_cell = ws.cell(row=2, column=qc)
    assert qc_cell.fill.start_color.rgb == RED
    assert "human review required" in qc_cell.value
    # Corrected sheet row should also be RED (not AMBER).
    cs = wb["Corrected"]
    assert cs.cell(row=2, column=1).fill.start_color.rgb == RED


def test_writer_warns_and_reds_empty_edits_at_med(tmp_path: Path):
    """Same guard for MED confidence — the rule is MED OR HIGH."""
    verdicts = _empty_edits_verdict(tmp_path, confidence="MED")
    out = tmp_path / "out.xlsx"
    proc = subprocess.run(
        [sys.executable, str(SCRIPT), "write", str(SAMPLE_XLSX), str(verdicts), str(out)],
        capture_output=True, text=True, check=True,
    )
    assert "NEEDS_EDITS at confidence MED but no edits supplied" in proc.stderr


def test_writer_does_not_warn_for_low_with_empty_edits(tmp_path: Path):
    """LOW + empty edits is the canonical 'external obstruction' path
    (construct mismatch, missing chart). No warning should fire."""
    verdicts = _empty_edits_verdict(tmp_path, confidence="LOW")
    out = tmp_path / "out.xlsx"
    proc = subprocess.run(
        [sys.executable, str(SCRIPT), "write", str(SAMPLE_XLSX), str(verdicts), str(out)],
        capture_output=True, text=True, check=True,
    )
    assert "no edits supplied" not in proc.stderr


def test_writer_does_not_warn_when_edits_are_present(tmp_path: Path):
    """A NEEDS_EDITS verdict with a non-empty edits array — no guard fire."""
    v = [
        {
            "row": 2,
            "status": "NEEDS_EDITS",
            "correctness_issue": "fake",
            "difficulty_issue": None,
            "edits": [
                {
                    "field": "option1",
                    "from": "<p>She don't likes coffee in the morning.</p>",
                    "to": "She doesn't likes coffee.",
                    "why": "test edit",
                }
            ],
            "confidence": "HIGH",
        }
    ]
    p = tmp_path / "with_edits.json"
    p.write_text(json.dumps(v))
    out = tmp_path / "out.xlsx"
    proc = subprocess.run(
        [sys.executable, str(SCRIPT), "write", str(SAMPLE_XLSX), str(p), str(out)],
        capture_output=True, text=True, check=True,
    )
    assert "no edits supplied" not in proc.stderr


# ---------------------------------------------------------------------------
# Legend sheet
# ---------------------------------------------------------------------------

def test_legend_sheet_has_four_color_swatches(written_workbook: Path):
    wb = openpyxl.load_workbook(written_workbook)
    ls = wb["QC Legend"]
    swatches = {
        2: ("GREEN", GREEN),
        3: ("AMBER", AMBER),
        4: ("DARK AMBER", DARK_AMBER),
        5: ("RED", RED),
    }
    for row_no, (label, fill) in swatches.items():
        cell = ls.cell(row=row_no, column=1)
        assert cell.value == label
        assert cell.fill.start_color.rgb == fill
