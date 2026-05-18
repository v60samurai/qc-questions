#!/usr/bin/env python3
"""
qc_xlsx.py — xlsx I/O for the qc-questions skill.

Two subcommands:

  read  <input.xlsx>
        Parses the xlsx, validates mandatory headers, strips HTML from
        content/option fields, and prints normalised JSON (one row per
        question) to stdout. The JSON for each question OMITS two fields:
          - correctOption — held back so the QC can solve blind.
          - difficulty    — held back so the QC can rate intrinsic complexity
                            blind, without being anchored to the marked band.
        Both audit targets are printed separately under "_keys" at the end
        for the QC to consume AFTER it has committed both a blind answer
        AND a blind difficulty rating per row.

  write <input.xlsx> <verdicts.json> <output.xlsx>
        Reads the original xlsx and a verdicts.json (a list of verdict
        objects keyed by row), and writes <output.xlsx> with the original
        sheet preserved plus appended qc_* columns.

Verdict object expected shape (lean, edit-centric — see REPORT_TEMPLATE.md):
{
  "row": <int>,
  "status": "ALIGNED" | "NEEDS_EDITS",
  "correctness_issue": "<one-line or null>",
  "difficulty_issue":  "<one-line or null>",
  "edits": [
    { "field": "<stem|option1..option6|correctOption|difficulty>",
      "from": "<exact current value>",
      "to":   "<exact replacement>",
      "why":  "<one-line>" }
  ],
  "confidence": "HIGH" | "MED" | "LOW"
}

The mandatory headers in the input xlsx (per skill spec) are:
  content, option1..option6, correctOption, questionType, subject, topics, difficulty
option5/option6 may be blank. Other columns are preserved untouched.

Dependencies: openpyxl (already used elsewhere in the user's environment).
HTML stripping uses only stdlib (html.parser).
"""

import argparse
import json
import sys
from html.parser import HTMLParser
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Color codes (Excel's standard Good/Neutral/Bad palette + accent)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")        # ALIGNED — no change
AMBER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")        # NEEDS_EDITS — apply edits
DARK_AMBER_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")   # changed cell accent
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")          # LOW confidence / escalate
BOLD_FONT = Font(bold=True)


MANDATORY_HEADERS = [
    "content",
    "option1",
    "option2",
    "option3",
    "option4",
    "option5",
    "option6",
    "correctOption",
    "questionType",
    "subject",
    "topics",
    "difficulty",
]

ORIGINAL_QC_COLUMN = "qc_status"           # Status column appended to the original sheet
ORIGINAL_QC_CHANGES_COLUMN = "qc_changes"  # Narrative audit column appended next to qc_status
CORRECTED_SHEET_NAME = "Corrected"         # New sheet containing post-edit rows
LEGEND_SHEET_NAME = "QC Legend"            # Small sheet documenting the color code
SUMMARY_SHEET_NAME = "QC Summary"          # Top-of-workbook tier-fit + blueprint-review observations (v0.4.0)

# Thresholds for the aggregate tier-fit observations rendered into the
# QC Summary sheet. Sourced from SKILL.md Rule 4-bis (blueprint-review)
# and REPORT_TEMPLATE.md Tier-Fit Observations. Each fires per (subject,
# topics) section, not bank-wide — a section-level miscalibration is the
# load-bearing signal a PM needs to see first.
OBSERVATION_THRESHOLDS = {
    "below_floor_pct":           40,   # > 40% items with blind < marked → over-tagged for the stated MQC
    "two_band_same_direction":   20,   # > 20% items 2-band gap in same direction → blueprint_review
    "formula_recall_pct":        50,   # > 50% HARD-tagged items with formula_recall_overrated → anti-pattern
    "construct_mismatch_pct":    10,   # > 10% items flagged as construct_mismatch
}

BAND_ORDER = ["EASY", "MEDIUM", "HARD"]

# Audit columns appended to the original sheet — excluded from the Corrected sheet.
ORIGINAL_AUDIT_COLUMNS = {ORIGINAL_QC_COLUMN, ORIGINAL_QC_CHANGES_COLUMN}

# Fields the LLM may target in `edits`. Edits to other fields are ignored
# (and IMMUTABLE_FIELDS specifically generate a warning to stderr).
EDITABLE_FIELDS = {
    "content",
    "stem",  # alias accepted; mapped to "content"
    "option1",
    "option2",
    "option3",
    "option4",
    "option5",
    "option6",
    "correctOption",
}

# These define what the question is supposed to BE — the PM-planned spec.
# The QC must edit `content`/`option*`/`correctOption` to align the item to
# these tags, never the other way around. Attempts to edit these are dropped
# and surfaced as a warning.
IMMUTABLE_FIELDS = {
    "subject",
    "topics",
    "difficulty",
}

# Cells that the bulk-upload platform expects to be HTML. Plain-text edits
# get wrapped in <p>…</p> so the corrected sheet is upload-ready.
HTML_WRAP_FIELDS = {
    "content",
    "option1",
    "option2",
    "option3",
    "option4",
    "option5",
    "option6",
}


def _normalize_field(field):
    return "content" if field == "stem" else field


def _format_qc_changes(verdict, applied_edits, *, effective_confidence=None):
    """Build the human-readable narrative for the qc_changes cell.

    Sections are separated by a blank line. `applied_edits` is the subset of
    `verdict["edits"]` that the writer actually wrote into the Corrected sheet
    (immutable-field edits that were dropped do not appear).

    `effective_confidence` lets the writer's auto-LOW guard override the raw
    verdict confidence — if the guard fired, we want the narrative to say
    "human review required" instead of the strengthen-distractors fallback.
    """
    if verdict.get("status") != "NEEDS_EDITS":
        return None

    conf = (
        effective_confidence
        if effective_confidence is not None
        else verdict.get("confidence")
    )

    sections = []
    correctness = verdict.get("correctness_issue")
    if correctness:
        sections.append(f"CORRECTNESS: {correctness}")
    difficulty = verdict.get("difficulty_issue")
    if difficulty:
        sections.append(f"DIFFICULTY: {difficulty}")

    if applied_edits:
        bullets = "\n".join(
            f"  • {_normalize_field(e.get('field'))}: {e.get('why', '')}"
            for e in applied_edits
        )
        sections.append("EDITS APPLIED:\n" + bullets)
    else:
        if conf == "LOW":
            sections.append("EDITS: none auto-applied — human review required")
        else:
            sections.append(
                "EDITS: none auto-applied — strengthen distractors per difficulty note"
            )

    return "\n\n".join(sections)


def _write_legend_sheet(wb):
    """Create (or replace) the QC Legend sheet documenting the colour code."""
    if LEGEND_SHEET_NAME in wb.sheetnames:
        del wb[LEGEND_SHEET_NAME]
    ls = wb.create_sheet(LEGEND_SHEET_NAME)

    rows = [
        ("Colour", "Where it appears",                                 "Meaning"),
        ("GREEN",      "qc_status cell (Original) · whole row (Corrected)",
                       "ALIGNED — correctness and difficulty both match the marked values. No edit needed."),
        ("AMBER",      "qc_status + qc_changes cells (Original) · whole row (Corrected)",
                       "NEEDS_EDITS — apply the edits shown in the Corrected sheet; qc_changes carries the narrative."),
        ("DARK AMBER", "individual cell in a NEEDS_EDITS row (Corrected)",
                       "This specific cell was edited. Compare with the same cell in the Original sheet to see the change."),
        ("RED",        "qc_status + qc_changes cells (Original) · whole row (Corrected)",
                       "NEEDS_EDITS with LOW confidence. Escalate to a human reviewer before applying."),
    ]

    fills_for_row = {
        1: None,
        2: GREEN_FILL,
        3: AMBER_FILL,
        4: DARK_AMBER_FILL,
        5: RED_FILL,
    }

    for r, row_vals in enumerate(rows, start=1):
        for c, val in enumerate(row_vals, start=1):
            cell = ls.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = BOLD_FONT
            else:
                # Colour-block the first column so the swatch is unambiguous.
                if c == 1:
                    cell.fill = fills_for_row[r]
                    cell.font = BOLD_FONT

    # Column widths for readability when the user opens the workbook.
    ls.column_dimensions["A"].width = 14
    ls.column_dimensions["B"].width = 48
    ls.column_dimensions["C"].width = 90

    # Trailing note row
    note_row = len(rows) + 2
    note = ls.cell(
        row=note_row,
        column=1,
        value=(
            "How to use: scan the qc_status column on the original sheet. For each amber or red row, "
            "read the qc_changes cell for the full narrative (correctness issue, difficulty issue, "
            "and every edit applied), then open the same row number on the 'Corrected' sheet — the "
            "cells highlighted in dark amber are the ones that changed. The Corrected sheet is the "
            "production-ready output: ALIGNED rows carry the original content verbatim so you can "
            "upload the sheet as-is (drop the fills first if your platform doesn't ignore them)."
        ),
    )
    note.font = Font(italic=True)
    ls.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)


def _maybe_html_wrap(field, value):
    if value is None:
        return value
    if field not in HTML_WRAP_FIELDS:
        return value
    s = str(value).strip()
    if s.startswith("<"):
        return s
    return f"<p>{s}</p>"


class _HTMLStripper(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self._parts: list[str] = []
        self._in_block = 0

    def handle_starttag(self, tag, attrs):
        if tag in ("p", "br", "li", "tr", "div"):
            self._parts.append("\n")
        elif tag in ("sup",):
            self._parts.append("^")
        elif tag in ("sub",):
            self._parts.append("_")

    def handle_endtag(self, tag):
        if tag in ("p", "li", "tr", "div"):
            self._parts.append("\n")

    def handle_data(self, data):
        self._parts.append(data)

    def text(self) -> str:
        joined = "".join(self._parts)
        # Collapse excessive whitespace but preserve line breaks loosely.
        lines = [ln.strip() for ln in joined.splitlines()]
        lines = [ln for ln in lines if ln]
        return "\n".join(lines).strip()


def strip_html(value) -> str:
    if value is None:
        return ""
    s = str(value)
    if "<" not in s and "&" not in s:
        return s.strip()
    parser = _HTMLStripper()
    try:
        parser.feed(s)
        parser.close()
    except Exception:
        return s.strip()
    return parser.text()


def _read_workbook(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    raw_headers = [c.value for c in ws[1]]
    headers = [h for h in raw_headers if h is not None]
    return wb, ws, raw_headers, headers


def _validate_headers(headers):
    missing = [h for h in MANDATORY_HEADERS if h not in headers]
    if missing:
        sys.stderr.write(
            "ERROR: input xlsx is missing mandatory headers: "
            + ", ".join(missing)
            + "\n"
            + "Mandatory headers: "
            + ", ".join(MANDATORY_HEADERS)
            + "\n"
        )
        sys.exit(2)


def cmd_read(args):
    in_path = Path(args.input).expanduser().resolve()
    if not in_path.exists():
        sys.stderr.write(f"ERROR: input file not found: {in_path}\n")
        sys.exit(2)

    _wb, ws, raw_headers, headers = _read_workbook(in_path)
    _validate_headers(headers)

    questions = []
    keys = []
    for excel_row, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        record = dict(zip(raw_headers, row, strict=False))
        # skip blank rows
        if not any(v not in (None, "") for v in record.values()):
            continue

        stripped = {
            "row": excel_row,
            "content": strip_html(record.get("content")),
            "option1": strip_html(record.get("option1")),
            "option2": strip_html(record.get("option2")),
            "option3": strip_html(record.get("option3")),
            "option4": strip_html(record.get("option4")),
            "option5": strip_html(record.get("option5")),
            "option6": strip_html(record.get("option6")),
            "questionType": record.get("questionType"),
            "subject": record.get("subject"),
            "topics": record.get("topics"),
            # NOTE: `difficulty` is held back in `_keys` — see comment below.
        }
        # carry through optional metadata for context (read-only),
        # but skip `difficulty` since it's the audit target being held back.
        for h in headers:
            if h in ("correctOption", "difficulty"):
                continue
            if h not in MANDATORY_HEADERS and h not in stripped:
                stripped[h] = record.get(h)

        # critical: hold back BOTH correctOption AND difficulty from the
        # question payload. correctOption is the answer-key audit target;
        # difficulty is the band-tag audit target. The QC must commit a
        # blind answer AND a blind difficulty rating per row before
        # opening _keys, otherwise both ratings are post-hoc rationalizations
        # of what was already visible.
        keys.append({
            "row": excel_row,
            "correctOption": record.get("correctOption"),
            "difficulty": record.get("difficulty"),
        })
        questions.append(stripped)

    out = {"questions": questions, "_keys": keys, "source": str(in_path)}
    json.dump(out, sys.stdout, indent=2, default=str)
    sys.stdout.write("\n")


def _band_gap(blind: str, marked: str):
    """Signed band gap: blind - marked. Negative = item easier than tagged for the stated MQC.
    Returns None if either input isn't a valid band."""
    if blind not in BAND_ORDER or marked not in BAND_ORDER:
        return None
    return BAND_ORDER.index(blind) - BAND_ORDER.index(marked)


def _compute_aggregate_observations(verdicts_list, ws, headers):
    """Group verdicts by (subject, topics) section and compute which tier-fit
    observations fire per section. Returns a dict with bank-level counts and
    a list of per-section observations sorted P0-first.

    Reads marked subject/topics/difficulty from the input sheet ws using the
    headers map. The verdict optionally carries `my_blind_difficulty` and
    `mistag_reason` (added in v0.4.0); when absent, the corresponding
    observations skip silently (no false fires).
    """
    try:
        sub_idx = headers.index("subject") + 1
        top_idx = headers.index("topics") + 1
        diff_idx = headers.index("difficulty") + 1
    except ValueError:
        return {"bank": {}, "sections": [], "observations": []}

    by_row = {v["row"]: v for v in verdicts_list}
    rows = []
    for excel_row in range(2, ws.max_row + 1):
        sub = ws.cell(row=excel_row, column=sub_idx).value
        top = ws.cell(row=excel_row, column=top_idx).value
        marked = ws.cell(row=excel_row, column=diff_idx).value
        if sub is None and top is None:
            continue
        v = by_row.get(excel_row, {})
        rows.append({
            "row": excel_row,
            "subject": sub,
            "topics": top,
            "marked": marked,
            "blind": v.get("my_blind_difficulty"),
            "status": v.get("status"),
            "confidence": v.get("confidence"),
            "mistag_reason": v.get("mistag_reason"),
            "correctness_issue": v.get("correctness_issue"),
            "difficulty_issue": v.get("difficulty_issue"),
            "construct_mismatch": bool(v.get("construct_mismatch")),
        })

    bank = {
        "total": len(rows),
        "aligned": sum(1 for r in rows if r["status"] == "ALIGNED"),
        "needs_edits": sum(1 for r in rows if r["status"] == "NEEDS_EDITS"),
        "low_confidence": sum(1 for r in rows if r["confidence"] == "LOW"),
    }

    # Group by section
    sections_map = {}
    for r in rows:
        key = (r["subject"] or "", r["topics"] or "")
        sections_map.setdefault(key, []).append(r)

    observations = []
    sections_summary = []
    for (sub, top), items in sections_map.items():
        n = len(items)
        # Below-floor: items with blind < marked (overrated for stated MQC), excluding rows without blind
        below_floor = [r for r in items if _band_gap(r["blind"], r["marked"]) is not None and _band_gap(r["blind"], r["marked"]) < 0]
        # Two-band same-direction
        two_band_over = [r for r in items if _band_gap(r["blind"], r["marked"]) == -2]
        two_band_under = [r for r in items if _band_gap(r["blind"], r["marked"]) == 2]
        # Formula-recall anti-pattern on HARD-tagged items
        hard_tagged = [r for r in items if r["marked"] == "HARD"]
        formula_recall = [r for r in hard_tagged if r["mistag_reason"] == "formula_recall_overrated"
                         or (r["difficulty_issue"] and "formula_recall_overrated" in r["difficulty_issue"])]
        # Construct mismatch
        construct = [r for r in items if r["construct_mismatch"]
                    or (r["correctness_issue"] and "construct mismatch" in r["correctness_issue"].lower())]

        pct_below = 100 * len(below_floor) / n if n else 0
        pct_2band_over = 100 * len(two_band_over) / n if n else 0
        pct_2band_under = 100 * len(two_band_under) / n if n else 0
        pct_formula = 100 * len(formula_recall) / len(hard_tagged) if hard_tagged else 0
        pct_construct = 100 * len(construct) / n if n else 0

        section_label = f"{sub} / {top}".strip(" /")
        sections_summary.append({
            "section": section_label, "n": n,
            "pct_below": pct_below, "pct_2band_over": pct_2band_over,
            "pct_2band_under": pct_2band_under, "pct_formula": pct_formula,
            "pct_construct": pct_construct,
            "rows_below": [r["row"] for r in below_floor],
            "rows_2band_over": [r["row"] for r in two_band_over],
            "rows_formula": [r["row"] for r in formula_recall],
        })

        # Emit observations (P0-ordered: blueprint_review first, then formula-recall, then below-floor, then construct)
        if pct_2band_over > OBSERVATION_THRESHOLDS["two_band_same_direction"]:
            observations.append({
                "severity": "P0", "kind": "blueprint_review", "section": section_label,
                "headline": f"Blueprint review — {section_label}: {len(two_band_over)}/{n} items "
                           f"({pct_2band_over:.0f}%) are 2-band-gapped OVERRATED for the stated MQC.",
                "detail": f"Same-direction 2-band gaps cluster in this section above the {OBSERVATION_THRESHOLDS['two_band_same_direction']}% threshold. "
                          f"This is a tag-calibration problem at the section level, not per-item drift. Affected rows: {sorted(set(r['row'] for r in two_band_over))}. "
                          f"Recommend a blueprint pass: either re-author the affected items to actually function at their marked bands, "
                          f"or relax the marked tags to match item behaviour for the audience.",
                "rows": sorted(set(r["row"] for r in two_band_over)),
            })
        if pct_2band_under > OBSERVATION_THRESHOLDS["two_band_same_direction"]:
            observations.append({
                "severity": "P0", "kind": "blueprint_review", "section": section_label,
                "headline": f"Blueprint review — {section_label}: {len(two_band_under)}/{n} items "
                           f"({pct_2band_under:.0f}%) are 2-band-gapped UNDERRATED for the stated MQC.",
                "detail": f"Items rated harder than tagged. Either author overestimates difficulty calibration or items are mistagged as EASY/MEDIUM. Affected rows: {sorted(set(r['row'] for r in two_band_under))}.",
                "rows": sorted(set(r["row"] for r in two_band_under)),
            })
        if pct_formula > OBSERVATION_THRESHOLDS["formula_recall_pct"]:
            observations.append({
                "severity": "P0", "kind": "formula_recall_anti_pattern", "section": section_label,
                "headline": f"Anti-pattern — {section_label}: {len(formula_recall)}/{len(hard_tagged)} HARD-tagged items "
                           f"({pct_formula:.0f}%) reduce to formula-recall + plug-in.",
                "detail": f"Formula recall is not HARD-band behaviour. Recommend re-authoring with multi-concept chaining, non-obvious decomposition, or interacting constraints. "
                          f"Affected rows: {sorted(set(r['row'] for r in formula_recall))}.",
                "rows": sorted(set(r["row"] for r in formula_recall)),
            })
        if pct_below > OBSERVATION_THRESHOLDS["below_floor_pct"]:
            observations.append({
                "severity": "P0", "kind": "section_below_floor", "section": section_label,
                "headline": f"Section composition — {section_label}: {len(below_floor)}/{n} items "
                           f"({pct_below:.0f}%) blind-rate easier than tagged for the stated MQC.",
                "detail": f"Section under-discriminates at the cut score. Information function is suboptimal. "
                          f"Recommend replacing {len(below_floor)} items with stems that function at their marked bands. Affected rows: {sorted(set(r['row'] for r in below_floor))}.",
                "rows": sorted(set(r["row"] for r in below_floor)),
            })
        if pct_construct > OBSERVATION_THRESHOLDS["construct_mismatch_pct"]:
            observations.append({
                "severity": "P0", "kind": "construct_drift", "section": section_label,
                "headline": f"Construct drift — {section_label}: {len(construct)}/{n} items "
                           f"({pct_construct:.0f}%) appear to test a different construct than tagged.",
                "detail": f"Recommend a tagging-taxonomy review before shipping. Affected rows: {sorted(set(r['row'] for r in construct))}.",
                "rows": sorted(set(r["row"] for r in construct)),
            })

    return {"bank": bank, "sections": sections_summary, "observations": observations}


def _write_summary_sheet(wb, aggregate):
    """Render the v0.4.0 QC Summary sheet at workbook index 0.

    Layout:
      Row 1: Title
      Row 2-3: Bank counts
      Rows 5+: Tier-fit observations (P0-first), one block per observation
      Below: Per-section breakdown table
    """
    if SUMMARY_SHEET_NAME in wb.sheetnames:
        del wb[SUMMARY_SHEET_NAME]
    s = wb.create_sheet(SUMMARY_SHEET_NAME, index=0)
    s.column_dimensions["A"].width = 22
    s.column_dimensions["B"].width = 100

    s["A1"] = "QC Summary — qc-questions v0.4.0"
    s["A1"].font = Font(bold=True, size=14)

    bank = aggregate["bank"]
    s["A3"] = "Bank counts:"
    s["A3"].font = BOLD_FONT
    s["B3"] = (
        f"total={bank.get('total', 0)}  |  ALIGNED={bank.get('aligned', 0)}  |  "
        f"NEEDS_EDITS={bank.get('needs_edits', 0)}  |  LOW-confidence={bank.get('low_confidence', 0)}"
    )

    row = 5
    observations = aggregate.get("observations", [])
    if observations:
        s.cell(row=row, column=1, value="Tier-fit observations (P0 — read these first):").font = Font(bold=True, size=12)
        row += 2
        for obs in observations:
            head_cell = s.cell(row=row, column=1, value=f"[{obs['severity']} {obs['kind']}]")
            head_cell.font = BOLD_FONT
            head_cell.fill = RED_FILL
            s.cell(row=row, column=2, value=obs["headline"]).font = BOLD_FONT
            s.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
            detail_cell = s.cell(row=row, column=2, value=obs["detail"])
            detail_cell.alignment = Alignment(wrap_text=True, vertical="top")
            s.row_dimensions[row].height = 60
            row += 2
    else:
        s.cell(row=row, column=1, value="Tier-fit observations:").font = BOLD_FONT
        s.cell(row=row, column=2, value="None fired. Bank-level calibration looks reasonable; review per-row verdicts.").font = Font(italic=True)
        row += 2

    # Per-section breakdown
    row += 1
    s.cell(row=row, column=1, value="Per-section breakdown:").font = Font(bold=True, size=12)
    row += 1
    headers_row = ["Section", "n", "% below floor", "% 2-band over", "% formula-recall (HARD-tagged)", "Rows below floor"]
    for col, h in enumerate(headers_row, start=1):
        c = s.cell(row=row, column=col, value=h)
        c.font = BOLD_FONT
    row += 1
    for sec in aggregate.get("sections", []):
        s.cell(row=row, column=1, value=sec["section"])
        s.cell(row=row, column=2, value=sec["n"])
        s.cell(row=row, column=3, value=f"{sec['pct_below']:.0f}%")
        s.cell(row=row, column=4, value=f"{sec['pct_2band_over']:.0f}%")
        s.cell(row=row, column=5, value=f"{sec['pct_formula']:.0f}%")
        s.cell(row=row, column=6, value=", ".join(str(r) for r in sec["rows_below"]) or "—")
        row += 1

    # Footnote
    row += 2
    s.cell(row=row, column=1, value="Thresholds:").font = BOLD_FONT
    s.cell(row=row, column=2, value=(
        f"blueprint_review fires > {OBSERVATION_THRESHOLDS['two_band_same_direction']}% same-direction 2-band gaps in section. "
        f"section_below_floor fires > {OBSERVATION_THRESHOLDS['below_floor_pct']}%. "
        f"formula_recall_anti_pattern fires > {OBSERVATION_THRESHOLDS['formula_recall_pct']}% of HARD-tagged items in section. "
        f"construct_drift fires > {OBSERVATION_THRESHOLDS['construct_mismatch_pct']}%."
    ))
    s.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    s.row_dimensions[row].height = 60


def cmd_write(args):
    in_path = Path(args.input).expanduser().resolve()
    verdicts_path = Path(args.verdicts).expanduser().resolve()
    out_path = Path(args.output).expanduser().resolve()

    if not in_path.exists():
        sys.stderr.write(f"ERROR: input file not found: {in_path}\n")
        sys.exit(2)
    if not verdicts_path.exists():
        sys.stderr.write(f"ERROR: verdicts file not found: {verdicts_path}\n")
        sys.exit(2)

    with verdicts_path.open() as f:
        verdicts_doc = json.load(f)
    verdicts_list = (
        verdicts_doc if isinstance(verdicts_doc, list) else verdicts_doc.get("verdicts", [])
    )
    by_row = {v["row"]: v for v in verdicts_list}

    wb, ws, raw_headers, headers = _read_workbook(in_path)
    _validate_headers(headers)

    # ---- 1. ORIGINAL sheet: append qc_status + qc_changes to the next truly
    # empty columns. Scan the entire sheet (not just the header row) so we
    # never overwrite data that happens to live past the last named header.
    refreshed_headers = [c.value for c in ws[1]]
    if ORIGINAL_QC_COLUMN in refreshed_headers:
        qc_status_col = refreshed_headers.index(ORIGINAL_QC_COLUMN) + 1
    else:
        max_col_with_data = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.column > max_col_with_data:
                    max_col_with_data = cell.column
        qc_status_col = max_col_with_data + 1
        header_cell = ws.cell(row=1, column=qc_status_col, value=ORIGINAL_QC_COLUMN)
        header_cell.font = BOLD_FONT

    # qc_changes lives immediately after qc_status. Detect existing placement
    # so re-runs don't drift; otherwise allocate the next column.
    refreshed_headers = [c.value for c in ws[1]]
    if ORIGINAL_QC_CHANGES_COLUMN in refreshed_headers:
        qc_changes_col = refreshed_headers.index(ORIGINAL_QC_CHANGES_COLUMN) + 1
    else:
        qc_changes_col = qc_status_col + 1
        header_cell = ws.cell(
            row=1, column=qc_changes_col, value=ORIGINAL_QC_CHANGES_COLUMN
        )
        header_cell.font = BOLD_FONT
    ws.column_dimensions[get_column_letter(qc_changes_col)].width = 80

    # ---- 2. CORRECTED sheet: rebuild from scratch.
    if CORRECTED_SHEET_NAME in wb.sheetnames:
        del wb[CORRECTED_SHEET_NAME]
    cs = wb.create_sheet(CORRECTED_SHEET_NAME)

    named_headers = [
        h for h in raw_headers if h is not None and h not in ORIGINAL_AUDIT_COLUMNS
    ]
    for i, h in enumerate(named_headers, start=1):
        hcell = cs.cell(row=1, column=i, value=h)
        hcell.font = BOLD_FONT

    # 3. Per row, apply verdict-driven behaviour:
    #    - ALIGNED      → write the original row content verbatim, fill GREEN.
    #                     The Corrected sheet is now the production-ready
    #                     source of truth for re-upload.
    #    - NEEDS_EDITS  → write the post-edit content, fill row AMBER
    #                     (or RED if LOW confidence). Cells that were actually
    #                     edited get DARK_AMBER + bold so the diff pops.
    #    Also write the qc_changes narrative back onto the original sheet.
    allow_retag = bool(getattr(args, "allow_retag", False))

    for excel_row in range(2, ws.max_row + 1):
        v = by_row.get(excel_row)

        # Start from the original row content for every status.
        record = {}
        for i, h in enumerate(raw_headers, start=1):
            if h is None or h in ORIGINAL_AUDIT_COLUMNS:
                continue
            record[h] = ws.cell(row=excel_row, column=i).value

        # Compute effective confidence (auto-LOW guard): a NEEDS_EDITS verdict
        # at MED/HIGH with an empty edits array is a QC regression — the
        # autonomous-by-default rule (SKILL.md rule 6) requires every
        # NEEDS_EDITS row to carry a concrete edit unless confidence is LOW
        # and the obstruction is external. Don't fail the write — surface the
        # gap with a stderr warning and paint the row RED so reviewers see it.
        status = v.get("status") if v else None
        effective_conf = v.get("confidence") if v else None
        if (
            v is not None
            and status == "NEEDS_EDITS"
            and effective_conf != "LOW"
            and not (v.get("edits") or [])
        ):
            sys.stderr.write(
                f"  ! row {excel_row}: NEEDS_EDITS at confidence {effective_conf} "
                f"but no edits supplied — QC should have prescribed an edit. "
                f"Treating as LOW for fill colour.\n"
            )
            effective_conf = "LOW"

        # Write qc_status on the original sheet.
        if status is not None:
            status_cell = ws.cell(row=excel_row, column=qc_status_col, value=status)
            if status == "ALIGNED":
                status_cell.fill = GREEN_FILL
            elif status == "NEEDS_EDITS":
                status_cell.fill = (
                    RED_FILL if effective_conf == "LOW" else AMBER_FILL
                )

        if v is None or status == "ALIGNED":
            # Corrected sheet: copy the original row verbatim, fill GREEN.
            for i, h in enumerate(named_headers, start=1):
                cell = cs.cell(row=excel_row, column=i, value=record.get(h))
                cell.fill = GREEN_FILL
            continue

        # NEEDS_EDITS: apply edits, tracking which actually landed.
        changed_fields = set()
        applied_edits = []
        for edit in v.get("edits") or []:
            field = _normalize_field(edit.get("field"))
            if field in IMMUTABLE_FIELDS:
                if not allow_retag:
                    sys.stderr.write(
                        f"  ! row {excel_row}: dropped edit targeting immutable field "
                        f"'{field}' (subject/topics/difficulty are the spec, not editable). "
                        f"Edit any of content/option1..option6/correctOption to align to "
                        f"the marked spec instead, or pass --allow-retag if you really "
                        f"intend to change tags (legacy-cleanup workflows only).\n"
                    )
                    continue
                if field not in record:
                    continue
                record[field] = edit.get("to")
                changed_fields.add(field)
                applied_edits.append(edit)
                continue
            if field not in EDITABLE_FIELDS or field not in record:
                continue
            record[field] = _maybe_html_wrap(field, edit.get("to"))
            changed_fields.add(field)
            applied_edits.append(edit)

        row_fill = RED_FILL if effective_conf == "LOW" else AMBER_FILL

        for i, h in enumerate(named_headers, start=1):
            cell = cs.cell(row=excel_row, column=i, value=record.get(h))
            if h in changed_fields:
                cell.fill = DARK_AMBER_FILL
                cell.font = BOLD_FONT
            else:
                cell.fill = row_fill

        # qc_changes narrative on the original sheet, fill mirrors qc_status.
        narrative = _format_qc_changes(
            v, applied_edits, effective_confidence=effective_conf
        )
        if narrative is not None:
            changes_cell = ws.cell(
                row=excel_row, column=qc_changes_col, value=narrative
            )
            changes_cell.fill = row_fill
            changes_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- 4. LEGEND sheet documenting the colour code.
    _write_legend_sheet(wb)

    # ---- 4. QC SUMMARY sheet (v0.4.0): compute aggregate observations and
    # render them at workbook index 0 so PMs see calibration issues BEFORE
    # the per-row verdicts. The aggregate logic uses optional verdict fields
    # (my_blind_difficulty, mistag_reason, construct_mismatch); when absent,
    # the sheet renders bank-level counts only — no false fires.
    aggregate = _compute_aggregate_observations(verdicts_list, ws, headers)
    _write_summary_sheet(wb, aggregate)

    # Reorder: QC Summary first, then Original → QC Legend → Corrected.
    sheet_order = [SUMMARY_SHEET_NAME]
    for name in wb.sheetnames:
        if name not in (SUMMARY_SHEET_NAME, LEGEND_SHEET_NAME, CORRECTED_SHEET_NAME):
            sheet_order.append(name)
    sheet_order.append(LEGEND_SHEET_NAME)
    sheet_order.append(CORRECTED_SHEET_NAME)
    wb._sheets = [wb[name] for name in sheet_order]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    sys.stderr.write(
        f"Wrote {out_path}\n"
        f"  - '{SUMMARY_SHEET_NAME}' sheet: bank counts + tier-fit observations "
        f"({len(aggregate['observations'])} fired) — read this FIRST\n"
        f"  - Original '{ws.title}': qc_status at column {qc_status_col}, "
        f"qc_changes at column {qc_changes_col}\n"
        f"  - '{LEGEND_SHEET_NAME}' sheet: colour-code legend\n"
        f"  - '{CORRECTED_SHEET_NAME}' sheet: ALIGNED rows populated verbatim+green "
        f"(upload-ready), NEEDS_EDITS rows show post-edit content (amber row, "
        f"dark-amber+bold on changed cells; red row if confidence=LOW or if "
        f"the auto-LOW guard fired for empty edits at MED/HIGH)\n"
    )


def main():
    p = argparse.ArgumentParser(description="QC xlsx I/O for the qc-questions skill")
    sub = p.add_subparsers(dest="cmd", required=True)

    pr = sub.add_parser("read", help="Parse xlsx and print normalised JSON")
    pr.add_argument("input")
    pr.set_defaults(func=cmd_read)

    pw = sub.add_parser("write", help="Write verdicts back into a new xlsx")
    pw.add_argument("input")
    pw.add_argument("verdicts")
    pw.add_argument("output")
    pw.add_argument(
        "--allow-retag",
        action="store_true",
        help=(
            "Allow verdict edits that target subject/topics/difficulty. "
            "OFF by default — the marked tags are the spec the question must match, "
            "and silently retagging would change bank composition. Enable only for "
            "legacy cleanup workflows where the tags themselves are known-bad."
        ),
    )
    pw.set_defaults(func=cmd_write)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
